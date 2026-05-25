"""Bank ledger Excel exporter — enriched with matched payment proof details."""

from __future__ import annotations

import io
from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")
_AMBER_FILL = PatternFill("solid", fgColor="FEF3C7")
_BLUE_FILL = PatternFill("solid", fgColor="DBEAFE")


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, 55)


def _dec(v: Any) -> Decimal:
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError):
        return Decimal("0")


def _build_match_index(job_blobs: dict[str, dict]) -> dict[str, dict]:
    """Map bank_entry.id → MatchResult dict across all job report blobs."""
    index: dict[str, dict] = {}
    for blob in job_blobs.values():
        for match in (blob or {}).get("matches", []):
            bank_entry = match.get("bank_entry") or {}
            entry_id = bank_entry.get("id")
            if entry_id:
                index[entry_id] = match
    return index


_MAIN_HEADERS = [
    "#",
    "Value Date",
    "Description",
    "Counterparty",
    "Amount (MYR)",
    "Currency",
    "Bank Reference",
    "Statement File",
    "Status",
    # Payment proof
    "Proof Date",
    "Payer",
    "Payee",
    "Original Amount",
    "Original Currency",
    "Invoice / Reference",
    # FX
    "Card FX Rate",
    "FX Rate (Invoice Date)",
    "FX Rate (Settlement Date)",
    "Amount MYR (Invoice Rate)",
    "Amount MYR (Settlement Rate)",
    # Reconciliation
    "Variance MYR",
    "Match Confidence",
    "Match Status",
    "Human Reviewed",
    "Review Notes",
    "Source Document",
    "Job ID",
]

_UNCLEARED_HEADERS = [
    "#",
    "Value Date",
    "Description",
    "Counterparty",
    "Amount (MYR)",
    "Currency",
    "Bank Reference",
    "Statement File",
    "Days Outstanding",
]


def _proof_cells(match: dict) -> list:
    nr = match.get("normalised_record") or {}
    payment = nr.get("payment") or {}
    return [
        payment.get("value_date", ""),
        payment.get("payer", ""),
        payment.get("payee", ""),
        payment.get("amount_original", ""),
        payment.get("currency", ""),
        payment.get("reference") or "",
        payment.get("card_fx_rate") or "",
        nr.get("fx_rate_invoice", ""),
        nr.get("fx_rate_settlement", ""),
        nr.get("amount_myr_at_invoice_rate", ""),
        nr.get("amount_myr_at_settlement_rate", ""),
        match.get("amount_variance_myr", ""),
        f"{match.get('confidence', 0):.2%}",
        match.get("status", ""),
        "Yes" if match.get("human_reviewed") else "No",
        match.get("review_notes") or "",
        (match.get("normalised_record") or {}).get("payment", {}).get("source_document") or "",
    ]


def _entry_row(i: int, entry, filename: str, match: dict | None) -> list:
    base: list[Any] = [
        i,
        entry.value_date,
        entry.description or "",
        entry.counterparty or "",
        float(entry.amount),
        entry.currency,
        entry.reference or "",
        filename,
        "Cleared" if entry.cleared else "Uncleared",
    ]
    base += _proof_cells(match) if match else [""] * 18
    return base


def render_ledger_excel(
    account,
    entries_with_filenames: list[tuple],
    job_blobs: dict[str, dict],
    date_from: date | None,
    date_to: date | None,
) -> bytes:
    wb = Workbook()
    today = date.today()
    match_index = _build_match_index(job_blobs)

    cleared_entries = [(e, f) for e, f in entries_with_filenames if e.cleared]
    uncleared_entries = [(e, f) for e, f in entries_with_filenames if not e.cleared]
    total_myr = sum(_dec(e.amount) for e, _ in entries_with_filenames)
    cleared_myr = sum(_dec(e.amount) for e, _ in cleared_entries)

    # ── Cover ──────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cover"
    ws.append(["ARIA Bank Ledger Export", ""])
    ws["A1"].font = Font(bold=True, size=14, color="1F2937")
    ws.append(["", ""])
    ws.append(["Account Name", account.name])
    ws.append(["Bank", account.bank_name])
    ws.append(["Account Number", account.account_number_masked])
    ws.append(["Currency", account.currency])
    ws.append(["", ""])
    ws.append(["Export Period", f"{date_from or 'All time'} to {date_to or today}"])
    ws.append(["Exported At", today.isoformat()])
    ws.append(["", ""])
    ws.append(["SUMMARY", ""])
    ws["A11"].font = Font(bold=True)
    ws.append(["Total Entries", len(entries_with_filenames)])
    ws.append(["Total Value (MYR)", float(total_myr)])
    ws.append(["Cleared Entries", len(cleared_entries)])
    ws.append(["Cleared Value (MYR)", float(cleared_myr)])
    ws.append(["Uncleared Entries", len(uncleared_entries)])
    ws.append(["Uncleared Value (MYR)", float(total_myr - cleared_myr)])
    ws.append(["", ""])
    ws.append(["LEGEND", ""])
    ws["A19"].font = Font(bold=True)

    legend_rows = [
        ("Green", "Cleared & auto-matched by ARIA"),
        ("Amber", "Cleared — confirmed by human reviewer"),
        ("No fill", "Uncleared — pending reconciliation"),
    ]
    fills = [_GREEN_FILL, _AMBER_FILL, None]
    for (label, desc), fill in zip(legend_rows, fills):
        ws.append([label, desc])
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 44

    # ── All Transactions ───────────────────────────────────────────────────────
    ws_all = wb.create_sheet("All Transactions")
    ws_all.append(_MAIN_HEADERS)
    _style_header(ws_all)
    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(_MAIN_HEADERS))}1"

    for i, (entry, filename) in enumerate(entries_with_filenames, 1):
        match = match_index.get(entry.id) if entry.cleared else None
        ws_all.append(_entry_row(i, entry, filename, match))
        if entry.cleared:
            fill = _AMBER_FILL if (match and match.get("human_reviewed")) else _GREEN_FILL
            for cell in ws_all[ws_all.max_row]:
                cell.fill = fill
    _autosize(ws_all)

    # ── Cleared ────────────────────────────────────────────────────────────────
    ws_c = wb.create_sheet("Cleared")
    ws_c.append(_MAIN_HEADERS)
    _style_header(ws_c)
    ws_c.freeze_panes = "A2"
    ws_c.auto_filter.ref = f"A1:{get_column_letter(len(_MAIN_HEADERS))}1"

    for i, (entry, filename) in enumerate(cleared_entries, 1):
        match = match_index.get(entry.id)
        ws_c.append(_entry_row(i, entry, filename, match))
        fill = _AMBER_FILL if (match and match.get("human_reviewed")) else _GREEN_FILL
        for cell in ws_c[ws_c.max_row]:
            cell.fill = fill
    _autosize(ws_c)

    # ── Uncleared ──────────────────────────────────────────────────────────────
    ws_u = wb.create_sheet("Uncleared")
    ws_u.append(_UNCLEARED_HEADERS)
    _style_header(ws_u)
    ws_u.freeze_panes = "A2"

    for i, (entry, filename) in enumerate(uncleared_entries, 1):
        days = (today - entry.value_date).days
        ws_u.append([
            i, entry.value_date, entry.description or "",
            entry.counterparty or "", float(entry.amount),
            entry.currency, entry.reference or "", filename, days,
        ])
    _autosize(ws_u)

    # ── FX Summary ─────────────────────────────────────────────────────────────
    ws_fx = wb.create_sheet("FX Summary")
    ws_fx.append([
        "Original Currency",
        "Cleared Entries",
        "Total Original Amount",
        "Avg Card FX Rate",
        "Avg Interbank Rate (Settlement)",
        "Total MYR (Settlement)",
        "Total Variance MYR",
    ])
    _style_header(ws_fx)

    by_ccy: dict[str, list[dict]] = defaultdict(list)
    for entry, _ in cleared_entries:
        match = match_index.get(entry.id)
        if not match:
            continue
        nr = match.get("normalised_record") or {}
        payment = nr.get("payment") or {}
        ccy = payment.get("currency") or entry.currency
        by_ccy[ccy].append({
            "orig": _dec(payment.get("amount_original", "0")),
            "card_fx": _dec(payment["card_fx_rate"]) if payment.get("card_fx_rate") else None,
            "fx_settle": _dec(nr.get("fx_rate_settlement", "0")),
            "myr": _dec(nr.get("amount_myr_at_settlement_rate", "0")),
            "variance": _dec(match.get("amount_variance_myr", "0")),
        })

    for ccy, recs in sorted(by_ccy.items()):
        n = len(recs)
        card_rates = [r["card_fx"] for r in recs if r["card_fx"] is not None]
        ws_fx.append([
            ccy,
            n,
            float(sum(r["orig"] for r in recs)),
            float(sum(card_rates) / len(card_rates)) if card_rates else "N/A",
            float(sum(r["fx_settle"] for r in recs) / n),
            float(sum(r["myr"] for r in recs)),
            float(sum(r["variance"] for r in recs)),
        ])
    _autosize(ws_fx)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
