"""Excel reconciliation report exporter."""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.schemas import AuditLogEntry, ReconciliationReport

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_AMBER_FILL = PatternFill("solid", fgColor="FEF3C7")
_GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")
_RED_FILL = PatternFill("solid", fgColor="FEE2E2")


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, 60)


def render_excel_report(
    report: ReconciliationReport,
    audit_log: list[AuditLogEntry] | None = None,
) -> bytes:
    """Render the 4-sheet Excel report described in the spec."""
    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    _style_header(ws)
    s = report.summary
    rows = [
        ("Job ID", str(report.job_id)),
        ("Generated at", report.generated_at.isoformat()),
        ("Base currency", report.base_currency),
        ("Total records", s.total_records),
        ("Matched", s.matched_count),
        ("Uncertain (review)", s.uncertain_count),
        ("Unmatched", s.unmatched_count),
        ("Total value (MYR)", str(s.total_value_myr)),
        ("Matched value (MYR)", str(s.matched_value_myr)),
        ("Total variance (MYR)", str(s.total_variance_myr)),
        ("Processing seconds", round(s.processing_seconds, 2)),
        ("Narrative", report.narrative),
    ]
    for row in rows:
        ws.append(list(row))
    _autosize(ws)

    # Matched
    ws = wb.create_sheet("Matched")
    ws.append([
        "Match ID",
        "Value date",
        "Payer",
        "Amount (original)",
        "Currency",
        "Amount (MYR)",
        "Bank entry date",
        "Bank entry amount (MYR)",
        "Bank description",
        "Bank reference",
        "Bank counterparty",
        "Confidence",
        "Variance (MYR)",
        "Reference",
        "Explanation",
    ])
    _style_header(ws)
    for m in report.matches:
        if m.status.value != "MATCHED":
            continue
        bank = m.bank_entry
        nr = m.normalised_record
        ws.append([
            str(m.id),
            nr.payment.value_date.isoformat(),
            nr.payment.payer,
            str(nr.payment.amount_original),
            nr.payment.currency,
            str(nr.amount_myr_at_settlement_rate),
            bank.value_date.isoformat() if bank else "",
            str(bank.amount) if bank else "",
            bank.description or "" if bank else "",
            bank.reference or "" if bank else "",
            bank.counterparty or "" if bank else "",
            f"{m.confidence:.2f}",
            str(m.amount_variance_myr),
            nr.payment.reference or "",
            m.variance_explanation,
        ])
    # Highlight rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = _GREEN_FILL
    _autosize(ws)

    # Exceptions
    ws = wb.create_sheet("Exceptions")
    ws.append([
        "Match ID",
        "Status",
        "Payer",
        "Amount (original)",
        "Currency",
        "Amount (MYR @ settlement)",
        "Tolerance low",
        "Tolerance high",
        "Confidence",
        "Reason",
        "Bank description",
        "Bank reference",
        "Bank counterparty",
    ])
    _style_header(ws)
    for m in report.matches:
        if m.status.value == "MATCHED":
            continue
        nr = m.normalised_record
        bank = m.bank_entry
        ws.append([
            str(m.id),
            m.status.value,
            nr.payment.payer,
            str(nr.payment.amount_original),
            nr.payment.currency,
            str(nr.amount_myr_at_settlement_rate),
            str(nr.tolerance_low),
            str(nr.tolerance_high),
            f"{m.confidence:.2f}",
            m.variance_explanation,
            bank.description if bank else "",
            bank.reference if bank else "",
            bank.counterparty if bank else "",
        ])
    for row in ws.iter_rows(min_row=2):
        fill = _AMBER_FILL if row[1].value == "UNCERTAIN" else _RED_FILL
        for cell in row:
            cell.fill = fill
    _autosize(ws)

    # Audit log
    ws = wb.create_sheet("Audit Log")
    ws.append([
        "Timestamp",
        "Agent",
        "Action",
        "Confidence",
        "Reasoning",
    ])
    _style_header(ws)
    for entry in audit_log or []:
        ws.append([
            entry.timestamp.isoformat(),
            entry.agent,
            entry.action,
            f"{entry.confidence:.2f}" if entry.confidence is not None else "",
            entry.reasoning,
        ])
    _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
