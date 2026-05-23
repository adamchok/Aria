"""Excel export — 4 sheets present with expected headers."""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from uuid import uuid4

import openpyxl

from app.models.enums import MatchStatus
from app.models.schemas import (
    AuditLogEntry,
    MatchResult,
    ReconciliationReport,
    ReconciliationSummary,
)
from app.services.excel_export import render_excel_report


def test_render_excel_has_four_sheets(normalised_record_usd, bank_entry_myr):
    matches = [
        MatchResult(
            normalised_record=normalised_record_usd,
            bank_entry=bank_entry_myr,
            confidence=0.9,
            status=MatchStatus.MATCHED,
            amount_variance_myr=Decimal("0.10"),
            variance_explanation="FX timing",
            reasoning_chain="...",
        ),
        MatchResult(
            normalised_record=normalised_record_usd,
            bank_entry=None,
            confidence=0.4,
            status=MatchStatus.UNMATCHED,
            amount_variance_myr=Decimal("0"),
            variance_explanation="no candidate",
            reasoning_chain="",
        ),
    ]
    report = ReconciliationReport(
        job_id=uuid4(),
        summary=ReconciliationSummary(
            total_records=2,
            matched_count=1,
            uncertain_count=0,
            unmatched_count=1,
            total_value_myr=Decimal("85.10"),
            matched_value_myr=Decimal("42.55"),
            total_variance_myr=Decimal("0.10"),
            processing_seconds=3.4,
        ),
        matches=matches,
        generated_at=datetime.utcnow(),
        base_currency="MYR",
        narrative="ARIA reconciled 1 of 2 records.",
    )
    audit = [
        AuditLogEntry(
            job_id=report.job_id,
            agent="ingestion",
            action="extract",
            confidence=0.95,
            reasoning="ok",
            timestamp=datetime.utcnow(),
        )
    ]
    data = render_excel_report(report, audit)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert {"Summary", "Matched", "Exceptions", "Audit Log"} <= set(wb.sheetnames)
    assert wb["Audit Log"].cell(row=1, column=1).value == "Timestamp"
